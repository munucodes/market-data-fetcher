"""
Compare two generated daily Excel outputs (values only).

Originally added to verify that refactoring the codebase into modules
did not change the daily Excel output. Can be reused for future changes.

Ignores formatting; compares cell values including intentional blanks.
"""

import sys
from openpyxl import load_workbook

def read_block(path, sheet_name=None):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    max_row = ws.max_row
    max_col = ws.max_column

    # Read all values (you can restrict if you know exact region)
    values = []
    for r in range(1, max_row + 1):
        row_vals = []
        for c in range(1, max_col + 1):
            row_vals.append(ws.cell(r, c).value)
        values.append(row_vals)
    return values

def main(a, b):
    va = read_block(a)
    vb = read_block(b)

    if len(va) != len(vb) or len(va[0]) != len(vb[0]):
        print("Different sheet dimensions.")
        print(f"{a}: {len(va)}x{len(va[0])}")
        print(f"{b}: {len(vb)}x{len(vb[0])}")
        sys.exit(2)

    diffs = []
    for r in range(len(va)):
        for c in range(len(va[0])):
            if va[r][c] != vb[r][c]:
                diffs.append((r+1, c+1, va[r][c], vb[r][c]))

    if diffs:
        print(f"Found {len(diffs)} differences. Showing first 20:")
        for d in diffs[:20]:
            print(f"Cell ({d[0]}, {d[1]}): {a}={d[2]!r} vs {b}={d[3]!r}")
        sys.exit(1)

    print("OK: outputs match (values).")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python compare_daily_outputs.py baseline.xlsx new.xlsx")
        sys.exit(2)
    main(sys.argv[1], sys.argv[2])
